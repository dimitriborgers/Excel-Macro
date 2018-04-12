import openpyxl
import read_excel
from calculate import compare_lists, process_content

#Create an excel workbook and set-up it's formatting
book = openpyxl.Workbook()
sheet1 = book.active
wrap_alignment = openpyxl.styles.Alignment(wrap_text=True,
                                            vertical='top',
                                            shrink_to_fit=True)
font = openpyxl.styles.Font(size=12)
book.save(read_excel.filename + "-Actual.xlsx")

#Creates the sheets with titles
def create_sheets():
    for i in read_excel.list1:
        book.create_sheet(title = i)
    book.remove_sheet(book.get_sheet_by_name('Sheet'))

#Adds on new calculation to new sheet
def iterate_sheets():
    create_sheets()

    #Create the lists that will keep track of the word counts, jaccard scores, average word count, 
    #and overlap between pairs
    word_count_list_1 = []
    word_count_list_2 = []
    jaccard_list = []
    average_count_list = []
    word_overlap_list = []
    word_list = []

    #Flip through each sheet, format the correct cells, and compare each pair within the sheet one by one
    for sheet in book.worksheets:
        #first format each row in the sheet
        for row in sheet['A1':'EE40']:
            for cell in row:
                cell.font = font
                cell.alignment = wrap_alignment

        #call the read function for read_excel import to retrieve all the data from the inputted file
        list2, list3, list4 = read_excel.read(sheet)

        #Write all the retrieved data on the workbook created at beginning of program
        write_sheets(sheet,list2,list3,list4)

        #temporary lists to keep track of the variables in this sheet alone
        column_number = 1
        word_count_list = []
        word_count_list_1_temp = []
        word_count_list_2_temp = []
        jaccard_list_temp = []
        word_overlap_list_temp = []
        word_list_temp = []

        #remove blanks in the list of sentences being compared
        list5 = [i for i in list4 if i != None]
        
        #Loop through the list of sentences, two at a time, and call the compare_lists function to retrieve
        #the jaccard scores, word overlap, and word counts
        for (i, k) in zip(list5[::2], list5[1::2]):
            jaccard_score, word_count_1, word_count_2, list_A, list_B, word_overlap = compare_lists(process_content(i), process_content(k))
            word_count_list.append(word_count_1)
            word_count_list.append(word_count_2)
            word_count_list_1_temp.append(word_count_1)
            word_count_list_2_temp.append(word_count_2)
            jaccard_list_temp.append(jaccard_score)
            word_overlap_list_temp.append(word_overlap)
            #word_list_temp = show_results(list_A, list_B, sheet, column_number)
        
        #Sum of all the numbers in each temporary list
        average_count = sum(word_count_list) / float(len(word_count_list))
        sum_count_1 = sum(word_count_list_1_temp)
        sum_count_2 = sum(word_count_list_2_temp)
        sum_jaccard = sum(jaccard_list_temp)
        word_overlap_count = sum(word_overlap_list_temp)

        #Add the summed up numbers in the non-temporary lists
        word_count_list_1.append(sum_count_1)
        word_count_list_2.append(sum_count_2)
        jaccard_list.append(sum_jaccard)
        average_count_list.append(average_count)
        word_overlap_list.append(word_overlap_count)
           
        #re-do the run through of the list of pairs and write down collected information on the new workbook
        for (i, k) in zip(list5[::2], list5[1::2]):
            jaccard_score, word_count_1, word_count_2, list_A, list_B, word_overlap = compare_lists(process_content(i), process_content(k))
               
            sheet.cell(row = 5, column = column_number).value = "Word Count 1"
            sheet.cell(row = 5, column = column_number + 1).value = word_count_1
            sheet.cell(row = 6, column = column_number).value = "Word Count 2"
            sheet.cell(row = 6, column = column_number + 1).value = word_count_2
            sheet.cell(row = 7, column = column_number).value = "Jaccard Score"
            sheet.cell(row = 7, column = column_number + 1).value = jaccard_score
            sheet.cell(row = 8, column = column_number).value = "Score/Avg Word Count"
            sheet.cell(row = 8, column = column_number + 1).value = float(jaccard_score / average_count)
            sheet.cell(row = 9, column = column_number).value = "Word Overlap"
            sheet.cell(row = 9, column = column_number + 1).value = word_overlap
            sheet.cell(row = 10, column = column_number + 1).value = "Sentence 1"
            sheet.cell(row = 10, column = column_number + 2).value = "Sentence 2"
            word_list_temp = show_results(list_A, list_B, sheet, column_number)
            column_number += 3
            word_list.extend(word_list_temp)
    #word_list = set(word_list)

    #Save the workbook and return the lists that hold all variables
    book.save(read_excel.filename + "-Actual.xlsx")
    return word_count_list_1, word_count_list_2, jaccard_list, average_count_list, word_overlap_list, word_list

#Writes down all information from existing Excel Sheet
def write_sheets(sheet,list2,list3,list4):
    for i in list2:
        if list2.index(i) < len(list2) + 1:
            sheet.cell(row = 1, column = list2.index(i) + 1).value = i

    for i in range(1,501):
        if i < len(list3) + 1:
            sheet.cell(row = 2, column = i).value = list3[i-1]

    for i in range(1,501):
        if i < len(list4) + 1:
            sheet.cell(row = 3, column = i).value = list4[i-1]

#Writes down each word that appeards in the pairs and show whether they are present in the left, right, or both 
#pairs by displaying a 1 if it is present and a 0 if it is not
def show_results(list1, list2, sheet, column_number):
    list_C = set.union(*[set(list1),set(list2)])
    list_D = list(set(list1) & set(list2))
    counter = 11
    for i in list_C:
        sheet.cell(row = counter, column = column_number).value = i
        sheet.cell(row = counter, column = column_number + 1).value = 0
        sheet.cell(row = counter, column = column_number + 2).value = 0
        if i in list1:
            sheet.cell(row = counter, column = column_number + 1).value = 1
        if i in list2: 
            sheet.cell(row = counter, column = column_number + 2).value = 1
        counter += 1
    return list_D
