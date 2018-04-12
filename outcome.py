import write_excel
from read_excel import list1
import openpyxl

#Open and edit the excel file whose name matches "filename"
filename = '(02.25.17)_overlap_skeleton.xlsx'
doc = openpyxl.load_workbook(filename, data_only = True)
sheet = doc.get_active_sheet()

#Create an excel workbook and set-up it's formatting
book = openpyxl.Workbook()
sheet1 = book.active
sheet2 = book.get_active_sheet()

#Run the write_excel function to collect the lists of data
word_count_list_1, word_count_list_2, jaccard_list, average_count_list, word_overlap_list, word_list = write_excel.iterate_sheets()

#Create indexes to go over cells in excel sheet
index_1 = 2
index_2 = 2
index_3 = 2
index_4 = 2
index_5 = 2
index_6 = 2
counter = 1

#Go through each list that we collected above and insert the data on each sheet and each row
for i in list1:
	sheet.cell(row=index_1,column=1).value = i
	index_1 += 1

for i in word_count_list_1:
	sheet.cell(row=index_2,column=3).value = i
	index_2 += 1

for i in word_count_list_2:
	sheet.cell(row=index_3,column=4).value = i
	index_3 += 1

for i in average_count_list:
	sheet.cell(row=index_4,column=6).value = i
	index_4 += 1

for i in jaccard_list:
	sheet.cell(row=index_5,column=7).value = i
	index_5 += 1

for i in word_overlap_list:
	sheet.cell(row=index_6,column=9).value = i
	index_6 += 1

doc.save(filename + "-Pseudo.xlsx")

for i in word_list:
	sheet2.cell(row=counter,column=1).value = i
	counter += 1

book.save("word_list.xlsx")
